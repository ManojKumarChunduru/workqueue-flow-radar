"""Render the daily ops packet: one Excel workbook plus a one-page PDF.

This is the artifact that replaces the manual morning report: per-queue
KPIs, the ping-pong hotlist, and the routing conflict table with the
hours and touches each conflict is costing. In shops running SSRS the
same content deploys as a scheduled report off these views; the file
packet ships first because it needs no serving surface (see the SSRS
note in the readme).
"""

from __future__ import annotations

import logging
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

log = logging.getLogger(__name__)

MAX_ROWS_PER_SHEET = 500

FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
FILL_ALERT = PatternFill("solid", fgColor="FFC7CE")
FONT_HEADER = Font(color="FFFFFF", bold=True, size=10)
FONT_TITLE = Font(bold=True, size=13)


def _write_frame(ws, df: pd.DataFrame, start_row: int = 1) -> None:
    for j, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row, column=j, value=str(col))
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(vertical="center")
    show = df.head(MAX_ROWS_PER_SHEET)
    for i, row in enumerate(show.itertuples(index=False), start=start_row + 1):
        for j, value in enumerate(row, start=1):
            if pd.isna(value):
                value = None
            elif isinstance(value, pd.Timestamp):
                value = value.to_pydatetime()
            ws.cell(row=i, column=j, value=value)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    ws.auto_filter.ref = (
        f"A{start_row}:{get_column_letter(len(df.columns))}{start_row + len(show)}"
    )
    for j, col in enumerate(df.columns, start=1):
        width = max(len(str(col)), *(len(str(v)) for v in show[col].head(200))) if len(show) else 12
        ws.column_dimensions[get_column_letter(j)].width = min(44, max(10, width + 2))
    if len(df) > MAX_ROWS_PER_SHEET:
        note = ws.cell(row=len(show) + start_row + 2, column=1,
                       value=f"Showing {MAX_ROWS_PER_SHEET} of {len(df)} rows. "
                             f"Full data: output/*.parquet")
        note.font = Font(italic=True)


def build_excel(
    out_dir: Path, kpis: pd.DataFrame, pingpong: pd.DataFrame,
    conflicts: pd.DataFrame, claim_stats: pd.DataFrame, headline: dict,
) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Workqueue Daily Ops Packet"
    ws["A1"].font = FONT_TITLE
    rows = [("Generated (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M"))]
    rows += list(headline.items())
    for i, (k, v) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=str(k)).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v if isinstance(v, (int, float)) else str(v))
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 22
    if headline.get("Ping-pong victims", 0):
        alert = ws.cell(row=len(rows) + 4, column=1,
                        value="Routing conflicts detected: see Routing Conflicts sheet")
        alert.fill = FILL_ALERT
        alert.font = Font(bold=True)

    wq_agg = (
        kpis.groupby("wq_id", as_index=False)
        .agg(
            visits=("visits_started", "sum"),
            touches=("touches", "sum"),
            defers=("defers", "sum"),
            resolved=("resolved", "sum"),
            avg_visit_hours=("avg_visit_hours", "mean"),
            p95_visit_hours=("p95_visit_hours", "max"),
        )
        .round({"avg_visit_hours": 2, "p95_visit_hours": 2})
        .sort_values("visits", ascending=False)
    )
    _write_frame(wb.create_sheet("WQ Overview"), wq_agg)
    _write_frame(wb.create_sheet("WQ Daily KPIs"), kpis.sort_values(["day", "wq_id"]))
    _write_frame(
        wb.create_sheet("Routing Conflicts"),
        conflicts if not conflicts.empty else pd.DataFrame({"status": ["no conflicts detected"]}),
    )
    _write_frame(
        wb.create_sheet("PingPong Hotlist"),
        pingpong.sort_values("total_hours", ascending=False)
        if not pingpong.empty
        else pd.DataFrame({"status": ["no ping-pong claims detected"]}),
    )

    out = out_dir / "wq-daily-ops-packet.xlsx"
    wb.save(out)
    log.info("excel packet written", extra={"path": str(out)})
    return out


def build_pdf(out_dir: Path, headline: dict, conflicts: pd.DataFrame) -> Path:
    out = out_dir / "wq-daily-ops-summary.pdf"
    body = ParagraphStyle("body", fontName="Helvetica", fontSize=9, leading=11.5,
                          textColor=colors.black, spaceAfter=4)
    h1 = ParagraphStyle("h1", fontName="Helvetica-Bold", fontSize=13, leading=15,
                        textColor=colors.black, spaceAfter=6)
    doc = SimpleDocTemplate(str(out), pagesize=letter, leftMargin=0.75 * inch,
                            rightMargin=0.75 * inch, topMargin=0.6 * inch,
                            bottomMargin=0.6 * inch, title="Workqueue daily ops summary")
    story = [Paragraph("Workqueue Daily Ops Summary", h1)]
    story.append(Paragraph(
        f"Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')} UTC. "
        "Full detail in the Excel packet.", body))
    story.append(Spacer(1, 6))

    rows = [["Metric", "Value"]] + [[str(k), str(v)] for k, v in headline.items()]
    t = Table(rows, colWidths=[3.2 * inch, 2.0 * inch])
    t.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.black),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(t)
    story.append(Spacer(1, 10))

    story.append(Paragraph("Routing conflicts (by victim claims)", h1))
    if conflicts.empty:
        story.append(Paragraph("No routing conflicts detected in this window.", body))
    else:
        cols = ["wq_x", "wq_y", "rule_into_x", "rule_into_y",
                "victim_claims", "victim_hours", "wasted_touches"]
        rows = [cols] + conflicts[cols].astype(str).values.tolist()
        ct = Table(rows, colWidths=[0.7 * inch, 0.7 * inch, 1.0 * inch, 1.0 * inch,
                                    1.0 * inch, 1.0 * inch, 1.1 * inch])
        ct.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 8.5),
            ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.black),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        story.append(ct)

    doc.build(story)
    log.info("pdf summary written", extra={"path": str(out)})
    return out


def build_packet(
    out_dir: str | Path, kpis: pd.DataFrame, pingpong: pd.DataFrame,
    conflicts: pd.DataFrame, claim_stats: pd.DataFrame,
) -> tuple[Path, Path]:
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    headline = {
        "Workqueues active": int(kpis["wq_id"].nunique()),
        "Claims worked": int(claim_stats["claim_id"].nunique()),
        "Claims resolved": int(kpis["resolved"].sum()),
        "First-pass yield": f"{claim_stats['first_pass'].mean():.1%}",
        "Median claim hours": round(float(claim_stats["total_hours"].median()), 1),
        "Ping-pong victims": int(len(pingpong)),
        "Hours lost to ping-pong": round(float(pingpong["total_hours"].sum()), 1)
        if not pingpong.empty else 0.0,
        "Conflicting rule pairs": int(len(conflicts)),
    }
    xlsx = build_excel(out_dir, kpis, pingpong, conflicts, claim_stats, headline)
    pdf = build_pdf(out_dir, headline, conflicts)
    return xlsx, pdf
